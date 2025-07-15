## backend/faktur/services/excel_exporter.py
# ==============================================================================

import os
import tempfile
from flask import send_file, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from models import PpnMasukan, PpnKeluaran

def generate_excel_export(db):
    try:

        wb = Workbook()
        ws = wb.active

        masukan = PpnMasukan.query.all()
        keluaran = PpnKeluaran.query.all()
        print("[DEBUG] Masukan:", len(masukan), "| Keluaran:", len(keluaran))

        def serialize_row(row):
            print(f"[DEBUG] Serialize: {row.id} | {row.no_faktur}")
            dpp_val = float(row.dpp)
            ppn_val = float(row.ppn)
            return [
                row.tanggal.strftime("%Y-%m-%d"),
                "PPN MASUKAN" if isinstance(row, PpnMasukan) else "PPN KELUARAN",
                row.keterangan,
                row.npwp_lawan_transaksi,
                row.nama_lawan_transaksi,
                row.no_faktur,
                dpp_val,
                ppn_val,
                dpp_val + ppn_val,  # 👈 Tambahan di akhir
            ]

        headers = [
            "Tanggal",
            "Jenis",
            "Keterangan",
            "NPWP Rekanan",
            "Nama Rekanan",
            "No Faktur",
            "DPP",
            "PPN",
            "Jumlah (DPP + PPN)",
        ]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data_rows = []
        for r in masukan + keluaran:
            try:
                data_rows.append(serialize_row(r))
            except Exception as e:
                print(f"[ERROR] Gagal serialize row ID={r.id}: {e}")

        # Load template
        BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        template_path = os.path.join(BASE_DIR, "templates", "rekap_template.xlsx")
        template_path = os.path.normpath(template_path)

        wb = load_workbook(template_path)
        ws = wb.active

        start_row = 2  # Anggap data mulai dari baris ke-3
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for idx, row_data in enumerate(data_rows, start=start_row):
            for col_index, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col_index, value=value)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                if isinstance(value, (int, float)) and col_index in [7, 8, 9]:
                    cell.number_format = "#,##0.00"

        # Simpan ke file sementara
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp.name)
        print("[DEBUG] Berhasil menyimpan file:", temp.name)
        return send_file(
            temp.name, as_attachment=True, download_name="rekap_pajak.xlsx"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500